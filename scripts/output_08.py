"""
output_08.py — Génération des reportings Excel
------------------------------------------------
Onglets générés :
  - PID & FR          : P&L consolidé PID + Financière RANMA
  - CELSIUS & VERTICAL: P&L consolidé CELSIUS + VERTICAL
  - Consolidé         : P&L groupe toutes entités
  - Bilan             : Bilan IFRS consolidé
  - Retraitements     : Récap éliminations intercos + IFRS 16

Inputs :
  - df_pl_final       : P&L après intercos (pcg_mapping_03 → agreger_pl)
  - df_bilan_mapped   : Bilan mappé (pcg_mapping_03 → agreger_bilan)
  - df_ca_pid         : Split CA PID par BU
  - df_ca_celsius     : Split CA CELSIUS par BU
  - df_cogs_pid       : Split COGS PID par BU
  - df_opex_rh        : Masse salariale OPEX par BU/Type
  - recap_pl          : Récap éliminations intercos P&L
  - recap_bs          : Récap éliminations intercos Bilan
  - ifrs16            : dict résultat ifrs16_07.run()
  - periode           : str YYYYMM
  - output_folder     : chemin de sortie
"""

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from config import (
    C_HEADER, C_SECTION, C_SUBTOTAL, C_TOTAL, C_ROW_ALT, C_WHITE, C_WARN,
    PL_STRUCTURE, REPORTING_GROUPS, IFRS16_ENTITIES,
)


# ── Styles (objets openpyxl) ──────────────────────────────────────────────────

THIN = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Sous-totaux calculés ──────────────────────────────────────────────────────

SUBTOTALS = {
    # Convention : charges stockées en négatif (agreger_pl applique * -1 sur classe 6 et 7)
    # → tous les items se somment directement, sans soustraction explicite
    "GROSS PROFIT"       : lambda d: d.get("Sales", 0) + d.get("B2C Revenue", 0) + d.get("B2B Revenue", 0) + d.get("COGS", 0),
    "CONTRIBUTION MARGIN": lambda d: (
        d.get("GROSS PROFIT", 0)
        + d.get("Staff costs (Operating)", 0)
        + d.get("Marketing costs", 0)
        + d.get("Freelance", 0)
        + d.get("Servers & softwares", 0)
    ),
    "EBITDA"             : lambda d: (
        d.get("CONTRIBUTION MARGIN", 0)
        + d.get("Staff costs (Non-op.)", 0)
        + d.get("Structure costs", 0)
        + d.get("Accommodation costs", 0)
        + d.get("Profit-sharing", 0)
        + d.get("Rents & charges", 0)
    ),
    "EBIT"               : lambda d: d.get("EBITDA", 0) + d.get("D&A on fixed assets", 0) + d.get("D&A - Milestones", 0) + d.get("D&A ROU (IFRS 16)", 0),
    "EBT"                : lambda d: d.get("EBIT", 0) + d.get("Financial income (loss)", 0),
    "NET INCOME"         : lambda d: d.get("EBT", 0) + d.get("Tax", 0),
}


# ── Helpers styling ───────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _style_cell(cell, row_type, col_idx, alt=False):
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left", vertical="center")

    if row_type == "section":
        cell.fill = _fill(C_SECTION)
        cell.font = _font(bold=True, color=C_WHITE, size=10)
    elif row_type == "subtotal":
        cell.fill = _fill(C_SUBTOTAL)
        cell.font = _font(bold=True, color=C_WHITE)
    elif row_type == "total":
        cell.fill = _fill(C_TOTAL)
        cell.font = _font(bold=True, color=C_WHITE)
    elif row_type == "spacer":
        cell.fill = _fill(C_WHITE)
    elif row_type == "item":
        cell.fill = _fill(C_ROW_ALT if alt else C_WHITE)
        cell.font = _font(bold=True)
    elif row_type == "detail":
        cell.fill = _fill(C_ROW_ALT if alt else C_WHITE)
        cell.font = _font(bold=False, size=9)
    else:
        cell.fill = _fill(C_ROW_ALT if alt else C_WHITE)
        cell.font = _font()

    if col_idx > 1 and row_type not in ("spacer", "section"):
        cell.number_format = '#,##0;[Red]-#,##0'


# ── Construction du dict de valeurs P&L pour un groupe d'entités ─────────────

def _build_pl_dict(entities, df_pl_final, df_opex_rh, ifrs16):
    """
    Retourne (d_flat, d_detail) pour un groupe d'entités.
      d_flat   : {category: montant_total}  — utilisé pour les sous-totaux
      d_detail : {category: {detail: montant}}  — utilisé pour le rendu à 2 niveaux
    """
    d_flat   = {}
    d_detail = {}

    # P&L standard (hors staff et rents)
    df = df_pl_final[df_pl_final["Entite"].isin(entities)].copy()

    for category, cat_grp in df.groupby("Mapping_PL_category"):
        d_flat[category] = d_flat.get(category, 0) + cat_grp["Mouvement"].sum()
        if category not in d_detail:
            d_detail[category] = {}
        for detail, det_grp in cat_grp.groupby("Mapping_PL_detail"):
            d_detail[category][detail] = (
                d_detail[category].get(detail, 0) + det_grp["Mouvement"].sum()
            )

    # Staff costs Operating / Non-operating
    if not df_opex_rh.empty:
        rh = df_opex_rh[df_opex_rh["Entite"].isin(entities)]
        op    = rh[rh["Type"].str.lower().str.contains("operat") & ~rh["Type"].str.lower().str.contains("non")]["Mouvement"].sum()
        nonop = rh[rh["Type"].str.lower().str.contains("non")]["Mouvement"].sum()
        d_flat["Staff costs (Operating)"]  = -op    # négatif : convention charges négatives
        d_flat["Staff costs (Non-op.)"]    = -nonop

    # IFRS 16 — neutralisation loyers + ROU D&A
    # Rents & charges est négatif (FEC classe 6 après * -1) ; loyers > 0 → on additionne pour annuler
    rou_total = 0
    for e in entities:
        if e in IFRS16_ENTITIES:
            key = e.lower()
            d_flat["Rents & charges"] = d_flat.get("Rents & charges", 0) + ifrs16[f"loyers_{key}"]
            rou_total += ifrs16[f"rou_{key}"]
    d_flat["D&A ROU (IFRS 16)"] = -rou_total  # charge D&A → négatif

    # Calcul des sous-totaux
    for ligne, fn in SUBTOTALS.items():
        d_flat[ligne] = fn(d_flat)

    return d_flat, d_detail


# ── Écriture d'un onglet P&L ──────────────────────────────────────────────────

def _write_pl_sheet(ws, title, col_groups, periode):
    """
    col_groups : liste de (label_colonne, d_flat, d_detail)
      d_flat   : {category: montant_total}
      d_detail : {category: {detail: montant}}
    """
    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(col_groups))
    tc = ws.cell(1, 1, f"{title} — {periode[:4]}/{periode[4:]}")
    tc.font = _font(bold=True, size=12, color=C_WHITE)
    tc.fill = _fill(C_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers colonnes
    ws.cell(2, 1, "").fill = _fill(C_HEADER)
    for ci, (lbl, _, _) in enumerate(col_groups, start=2):
        c = ws.cell(2, ci, lbl)
        c.fill = _fill(C_HEADER)
        c.font = _font(bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER_THIN
    ws.row_dimensions[2].height = 18

    # Lignes P&L (row dynamique pour absorber les lignes de détail)
    row = 3
    alt  = False

    for ligne, row_type in PL_STRUCTURE:
        ws.row_dimensions[row].height = 16 if row_type != "spacer" else 6
        label_cell = ws.cell(row, 1, ligne if row_type != "spacer" else "")
        _style_cell(label_cell, row_type, 1, alt)

        for ci, (_, d_flat, _) in enumerate(col_groups, start=2):
            val = d_flat.get(ligne, 0) if row_type not in ("section", "spacer") else ""
            c = ws.cell(row, ci, val if val != 0 or row_type in ("subtotal", "total") else "")
            _style_cell(c, row_type, ci, alt)

        row += 1

        if row_type == "item":
            alt = not alt

            # Détail : sous-lignes de la catégorie
            all_details = []
            seen = set()
            for _, _, d_detail in col_groups:
                for det in d_detail.get(ligne, {}):
                    if det not in seen:
                        all_details.append(det)
                        seen.add(det)

            for detail in all_details:
                ws.row_dimensions[row].height = 14
                lc = ws.cell(row, 1, f"   {detail}")
                _style_cell(lc, "detail", 1, alt)

                for ci, (_, _, d_detail) in enumerate(col_groups, start=2):
                    val = d_detail.get(ligne, {}).get(detail, 0)
                    c = ws.cell(row, ci, val if val != 0 else "")
                    _style_cell(c, "detail", ci, alt)

                row += 1
                alt = not alt

        else:
            alt = False

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 35
    for ci in range(2, 2 + len(col_groups)):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ── Onglet Bilan ──────────────────────────────────────────────────────────────

def _write_bilan_sheet(ws, df_bilan_mapped, periode):
    headers = ["Ligne Bilan"] + ["FR", "PID", "CELSIUS", "VERTICAL", "CONSOLIDÉ"]
    entites = ["FR", "PID", "CELSIUS", "VERTICAL"]

    # Pivot
    pivot = df_bilan_mapped.pivot_table(
        index="Mapping_BS_detail", columns="Entite", values="Solde", aggfunc="sum"
    ).reindex(columns=entites).fillna(0)
    pivot["CONSOLIDÉ"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(1, 1, f"Bilan IFRS — {periode[:4]}/{periode[4:]}")
    tc.font = _font(bold=True, size=12, color=C_WHITE)
    tc.fill = _fill(C_HEADER)
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.fill = _fill(C_HEADER)
        c.font = _font(bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal="right" if ci > 1 else "left")
        c.border = BORDER_THIN

    # Données
    for ri, row in enumerate(pivot.itertuples(index=False), start=3):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, start=1):
            c = ws.cell(ri, ci, val)
            c.fill = _fill(C_ROW_ALT if alt else C_WHITE)
            c.font = _font()
            c.border = BORDER_THIN
            if ci > 1:
                c.number_format = '#,##0;[Red]-#,##0'
                c.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 35
    for ci in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# ── Onglet Retraitements ──────────────────────────────────────────────────────

def _write_retraitements_sheet(ws, recap_pl, recap_bs, ifrs16, periode):
    row = 1

    def _section_title(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row, 1, title)
        c.fill = _fill(C_SECTION)
        c.font = _font(bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal="left")
        ws.row_dimensions[row].height = 18
        row += 1

    def _write_df(df, col_names):
        nonlocal row
        for ci, h in enumerate(col_names, 1):
            c = ws.cell(row, ci, h)
            c.fill = _fill(C_HEADER)
            c.font = _font(bold=True, color=C_WHITE)
            c.border = BORDER_THIN
        row += 1
        for _, r in df.iterrows():
            alt = row % 2 == 0
            for ci, val in enumerate(r, 1):
                c = ws.cell(row, ci, val)
                c.fill = _fill(C_ROW_ALT if alt else C_WHITE)
                c.font = _font()
                c.border = BORDER_THIN
                # Alerte écart
                if isinstance(val, float) and col_names[ci-1] == "Ecart" and abs(val) > 0.01:
                    c.font = _font(bold=True, color=C_WARN)
            row += 1
        row += 1

    # Éliminations P&L
    _section_title(f"Éliminations intercos P&L — {periode[:4]}/{periode[4:]}")
    if not recap_pl.empty:
        _write_df(recap_pl, list(recap_pl.columns))

    # Éliminations BS
    _section_title(f"Éliminations intercos Bilan — {periode[:4]}/{periode[4:]}")
    if not recap_bs.empty:
        _write_df(recap_bs, list(recap_bs.columns))

    # IFRS 16
    _section_title("Retraitement IFRS 16")
    ifrs_data = ifrs16["df_ifrs16"]
    _write_df(ifrs_data, list(ifrs_data.columns))

    ws.column_dimensions["A"].width = 30
    for ci in range(2, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ── Onglet Détail P&L FEC ─────────────────────────────────────────────────────

def _write_pl_detail_sheet(ws, df_pl_elimine, periode):
    """Comptes FEC individuels avec leur mapping P&L, groupés par (Entité, Mapping_PL_detail)."""
    df = df_pl_elimine[
        df_pl_elimine["ClasseCompte"].isin(["6", "7"]) &
        df_pl_elimine["Mapping_PL_detail"].notna()
    ].copy()

    # Convention P&L : même inversion que agreger_pl
    df["Mouvement_PL"] = df["Mouvement"] * -1
    df = df.sort_values(["Entite", "Mapping_PL_detail", "CompteNum"]).reset_index(drop=True)

    NB_COLS = 5  # Entité | N° Compte | Libellé | Mapping P&L | Mouvement

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NB_COLS)
    tc = ws.cell(1, 1, f"Détail P&L FEC — {periode[:4]}/{periode[4:]}")
    tc.font = _font(bold=True, size=12, color=C_WHITE)
    tc.fill = _fill(C_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, h in enumerate(["Entité", "N° Compte", "Libellé Compte", "Mapping P&L", "Mouvement"], 1):
        c = ws.cell(2, ci, h)
        c.fill = _fill(C_HEADER)
        c.font = _font(bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal="right" if ci == NB_COLS else "left")
        c.border = BORDER_THIN
    ws.row_dimensions[2].height = 18

    row = 3
    for (entite, mapping), grp in df.groupby(["Entite", "Mapping_PL_detail"], sort=True):
        # Ligne section : label + sous-total
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NB_COLS - 1)
        ch = ws.cell(row, 1, f"{entite}  ·  {mapping}")
        ch.fill = _fill(C_SECTION)
        ch.font = _font(bold=True, color=C_WHITE)
        ch.alignment = Alignment(horizontal="left", vertical="center")
        ch.border = BORDER_THIN
        cs = ws.cell(row, NB_COLS, grp["Mouvement_PL"].sum())
        cs.fill = _fill(C_SECTION)
        cs.font = _font(bold=True, color=C_WHITE)
        cs.number_format = '#,##0;[Red]-#,##0'
        cs.alignment = Alignment(horizontal="right")
        cs.border = BORDER_THIN
        ws.row_dimensions[row].height = 16
        row += 1

        # Lignes comptes
        for i, (_, r) in enumerate(grp.iterrows()):
            alt = i % 2 == 0
            for ci, val in enumerate([r["Entite"], r["CompteNum"], r["CompteLib"], r["Mapping_PL_detail"], r["Mouvement_PL"]], 1):
                c = ws.cell(row, ci, val)
                c.fill = _fill(C_ROW_ALT if alt else C_WHITE)
                c.font = _font()
                c.border = BORDER_THIN
                if ci == NB_COLS:
                    c.number_format = '#,##0;[Red]-#,##0'
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left")
            ws.row_dimensions[row].height = 15
            row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16


# ── Point d'entrée principal ──────────────────────────────────────────────────

def run(
    df_pl_final,
    df_pl_elimine,
    df_bilan_mapped,
    df_opex_rh,
    recap_pl,
    recap_bs,
    ifrs16,
    periode,
    output_folder="data/output",
):
    Path(output_folder).mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # Supprime la feuille vide par défaut

    # ── Onglets P&L ───────────────────────────────────────────────────────────
    for sheet_name, entities in REPORTING_GROUPS.items():
        ws = wb.create_sheet(sheet_name)

        # Colonnes = une par entité + total groupe
        col_groups = []
        for e in entities:
            d_flat, d_detail = _build_pl_dict([e], df_pl_final, df_opex_rh, ifrs16)
            col_groups.append((e, d_flat, d_detail))

        d_flat_total, d_detail_total = _build_pl_dict(entities, df_pl_final, df_opex_rh, ifrs16)
        col_groups.append(("TOTAL", d_flat_total, d_detail_total))

        _write_pl_sheet(ws, sheet_name, col_groups, periode)
        print(f"[output_08] Onglet '{sheet_name}' généré")

    # ── Bilan ─────────────────────────────────────────────────────────────────
    ws_bilan = wb.create_sheet("Bilan")
    _write_bilan_sheet(ws_bilan, df_bilan_mapped, periode)
    print("[output_08] Onglet 'Bilan' généré")

    # ── Retraitements ─────────────────────────────────────────────────────────
    ws_ret = wb.create_sheet("Retraitements")
    _write_retraitements_sheet(ws_ret, recap_pl, recap_bs, ifrs16, periode)
    print("[output_08] Onglet 'Retraitements' généré")

    # ── Détail P&L FEC ────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Détail P&L FEC")
    _write_pl_detail_sheet(ws_detail, df_pl_elimine, periode)
    print("[output_08] Onglet 'Détail P&L FEC' généré")

    # ── Sauvegarde ────────────────────────────────────────────────────────────
    filename = f"reporting_{periode}.xlsx"
    filepath = Path(output_folder) / filename
    wb.save(filepath)
    print(f"\n[output_08] ✅ Fichier généré : {filepath}")
    return str(filepath)
